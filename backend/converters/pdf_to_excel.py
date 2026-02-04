# -*- coding: utf-8 -*-
"""
PDF to Excel Converter
Camelot ile PDF tablolarını Excel'e dönüştürme
"""

import io
import os
import tempfile
from typing import List, Dict, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from core.pdf_reader import PDFReader
from translators.gemini_translator import get_translator


class PDFToExcelConverter:
    """
    PDF'ten Excel'e dönüştürücü
    Tabloları algılar ve Excel formatında çıktı verir
    """

    def __init__(self):
        self.translator = get_translator()

    def convert(self, pdf_bytes: bytes, source_lang: str = "auto",
               target_lang: str = "tr", translate: bool = False,
               extract_all: bool = True) -> bytes:
        """
        PDF'i Excel'e dönüştür

        Args:
            pdf_bytes: PDF bayt verisi
            source_lang: Kaynak dil
            target_lang: Hedef dil
            translate: Çeviri yap
            extract_all: Tüm metni çıkar (tablo olmayanlar için)

        Returns:
            bytes: XLSX bayt verisi
        """
        # PDF'i oku
        with PDFReader(pdf_bytes=pdf_bytes) as reader:
            all_tables = []
            all_text = []

            for page_num in range(len(reader)):
                # Tabloları çıkar (basit yaklaşım)
                tables = self._extract_tables_from_page(reader, page_num)
                all_tables.extend(tables)

                # Tüm metni çıkar
                if extract_all:
                    text = reader.get_full_text(page_num)
                    if text.strip():
                        all_text.append({
                            "page": page_num + 1,
                            "text": text
                        })

            # Excel oluştur
            output = io.BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Tabloları yaz
                for i, table in enumerate(all_tables):
                    df = pd.DataFrame(table)

                    if translate:
                        # Hücreleri çevir
                        df = self._translate_dataframe(df, source_lang, target_lang)

                    sheet_name = f"Tablo_{i+1}"
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Metin sayfası
                if extract_all and all_text:
                    text_df = pd.DataFrame(all_text)
                    if translate:
                        text_df["text"] = text_df["text"].apply(
                            lambda x: self._translate_text(x, source_lang, target_lang)
                        )
                    text_df.to_excel(writer, sheet_name="Metin", index=False)

            return output.getvalue()

    def _extract_tables_from_page(self, reader: PDFReader, page_num: int) -> List[List[List[str]]]:
        """
        Sayfadan tabloları çıkar

        Args:
            reader: PDFReader
            page_num: Sayfa numarası

        Returns:
            List of tables (each table is a 2D list)
        """
        page = reader.get_page(page_num)

        # Metin bloklarını al
        blocks = page.get_text("dict")["blocks"]

        # Basit tablo tespiti
        tables = []
        current_table = []
        current_row = []

        # Y koordinatlarına göre blokları sırala
        text_blocks = []
        for block in blocks:
            if block["type"] == 0:  # Metin
                text = ""
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        text += span["text"] + " "
                if text.strip():
                    text_blocks.append({
                        "text": text.strip(),
                        "bbox": block["bbox"]
                    })

        # Y'ye göre sırala (yukarıdan aşağı)
        text_blocks.sort(key=lambda b: b["bbox"][1])

        # Tablo yapısı tespiti (basit)
        # Aynı y seviyesindeki blokları satır olarak kabul et
        if text_blocks:
            current_y = text_blocks[0]["bbox"][1]
            current_row = [text_blocks[0]["text"]]

            for block in text_blocks[1:]:
                y = block["bbox"][1]

                # Y toleransı (10px)
                if abs(y - current_y) < 10:
                    current_row.append(block["text"])
                else:
                    # Yeni satır
                    if current_row:
                        current_table.append(current_row)
                    current_row = [block["text"]]
                    current_y = y

            if current_row:
                current_table.append(current_row)

            if current_table:
                tables.append(current_table)

        return tables

    def _translate_dataframe(self, df: pd.DataFrame, source_lang: str,
                            target_lang: str) -> pd.DataFrame:
        """DataFrame'i çevir"""
        translated_df = df.copy()

        for col in translated_df.columns:
            translated_df[col] = translated_df[col].apply(
                lambda x: self._translate_text(str(x), source_lang, target_lang)
            )

        return translated_df

    def _translate_text(self, text: str, source_lang: str, target_lang: str) -> str:
        """Metni çevir"""
        if not text or text.strip() == "" or text == "nan":
            return text

        try:
            result = self.translator.translate(text, target_lang, source_lang)
            return result.text if result.success else text
        except:
            return text


class AdvancedPDFToExcelConverter:
    """
    Gelişmiş PDF to Excel converter
    Camelot entegrasyonu ile daha iyi tablo algılama
    """

    def __init__(self):
        self.translator = get_translator()

    def convert_with_camelot(self, pdf_bytes: bytes, source_lang: str = "auto",
                            target_lang: str = "tr", translate: bool = False) -> bytes:
        """
        Camelot ile PDF'ten Excel'e dönüştür

        Args:
            pdf_bytes: PDF bayt verisi
            source_lang: Kaynak dil
            target_lang: Hedef dil
            translate: Çeviri yap

        Returns:
            bytes: XLSX bayt verisi
        """
        # Geçici dosya oluştur
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as pdf_file:
            pdf_file.write(pdf_bytes)
            pdf_path = pdf_file.name

        try:
            import camelot

            # Tabloları çıkar
            tables = camelot.read_pdf(pdf_path, pages='all')

            output = io.BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                for i, table in enumerate(tables):
                    df = table.df

                    if translate:
                        df = self._translate_dataframe(df, source_lang, target_lang)

                    sheet_name = f"Tablo_{i+1}"
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            return output.getvalue()

        except ImportError:
            print("⚠️ Camelot kurulu değil, varsayılan converter kullanılıyor")
            converter = PDFToExcelConverter()
            return converter.convert(pdf_bytes, source_lang, target_lang, translate)

        except Exception as e:
            print(f"⚠️ Camelot hatası: {e}, varsayılan converter kullanılıyor")
            converter = PDFToExcelConverter()
            return converter.convert(pdf_bytes, source_lang, target_lang, translate)

        finally:
            if os.path.exists(pdf_path):
                os.unlink(pdf_path)

    def _translate_dataframe(self, df: pd.DataFrame, source_lang: str,
                            target_lang: str) -> pd.DataFrame:
        """DataFrame'i çevir"""
        translated_df = df.copy()

        for col in translated_df.columns:
            translated_df[col] = translated_df[col].apply(
                lambda x: self._translate_text(str(x), source_lang, target_lang)
            )

        return translated_df

    def _translate_text(self, text: str, source_lang: str, target_lang: str) -> str:
        """Metni çevir"""
        if not text or text.strip() == "" or text == "nan":
            return text

        try:
            result = self.translator.translate(text, target_lang, source_lang)
            return result.text if result.success else text
        except:
            return text


class StyledExcelConverter:
    """
    Stilli Excel çıktısı
    Renkler, kenarlıklar, hizalamalar
    """

    def __init__(self):
        self.translator = get_translator()

    def convert_styled(self, pdf_bytes: bytes, source_lang: str = "auto",
                      target_lang: str = "tr", translate: bool = False) -> bytes:
        """
        Stilli Excel oluştur

        Args:
            pdf_bytes: PDF bayt verisi
            source_lang: Kaynak dil
            target_lang: Hedef dil
            translate: Çeviri yap

        Returns:
            bytes: XLSX bayt verisi
        """
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Tabloları çıkar
        converter = PDFToExcelConverter()
        raw_output = converter.convert(pdf_bytes, source_lang, target_lang,
                                      translate=False)

        # Excel'i yükle ve stilleri uygula
        wb = Workbook(io.BytesIO(raw_output))

        # Her sayfa için
        for sheet in wb.worksheets:
            # Başlık satırını stillendir
            for cell in sheet[1]:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="CCCCCC",
                                       end_color="CCCCCC",
                                       fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Kenarlıklar ekle
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True)

        # Çıktı
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


def convert_pdf_to_excel(pdf_bytes: bytes, translate: bool = False,
                        source_lang: str = "auto", target_lang: str = "tr",
                        use_camelot: bool = True) -> bytes:
    """
    PDF'ten Excel'e dönüşüm (kolay fonksiyon)

    Args:
        pdf_bytes: PDF bayt verisi
        translate: Çeviri yap
        source_lang: Kaynak dil
        target_lang: Hedef dil
        use_camelot: Camelot kullan (daha iyi tablo algılama)

    Returns:
        bytes: XLSX bayt verisi
    """
    if use_camelot:
        converter = AdvancedPDFToExcelConverter()
        return converter.convert_with_camelot(pdf_bytes, source_lang, target_lang, translate)
    else:
        converter = PDFToExcelConverter()
        return converter.convert(pdf_bytes, source_lang, target_lang, translate)
